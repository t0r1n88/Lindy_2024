"""
Склонение ФИО по падежам
"""
import openpyxl
import pandas as pd
from tkinter import messagebox
import re
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from pytrovich.detector import PetrovichGenderDetector
from pytrovich.enums import NamePart, Gender, Case
from pytrovich.maker import PetrovichDeclinationMaker
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None
import logging
logging.basicConfig(
    level=logging.WARNING,
    filename="error.log",
    filemode='w',
    # чтобы файл лога перезаписывался  при каждом запуске.Чтобы избежать больших простыней. По умолчанию идет 'a'
    format="%(asctime)s - %(module)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
    datefmt='%H:%M:%S',
)

class NotFIOPart(Exception):
    """
    исключение для проверки наличия колонок Фамилия,Имя, Отчество
    """
    pass

class FIOTogether(Exception):
    """
    исключение для проверки одновременного наличия колонок ФИО и колонок Фамилия,Имя,Отчество
    """
    pass



def capitalize_double_name(word):
    """
    Функция для того чтобы в двойных именах и фамилиях, отчествах вторая часть была также с большой буквы
    """
    lst_word = word.split('-')  # сплитим по дефису
    if len(lst_word) == 1:  # если длина списка равна 1 то это не двойная фамилия и просто возвращаем слово

        return word
    elif len(lst_word) == 2:
        first_word = lst_word[0].capitalize()  # делаем первую букву слова заглавной а остальные строчными
        second_word = lst_word[1].capitalize()
        return f'{first_word}-{second_word}'
    else:
        return 'Не удалось просклонять'


def case_lastname(maker, lastname, gender, case: Case):
    """
    Функция для обработки и склонения фамилии. Это нужно для обработки случаев двойной фамилии
    """

    lst_lastname = lastname.split('-')  # сплитим по дефису

    if len(lst_lastname) == 1:  # если длина списка равна 1 то это не двойная фамилия и просто обрабатываем слово
        case_result_lastname = maker.make(NamePart.LASTNAME, gender, case, lastname)
        return case_result_lastname
    elif len(lst_lastname) == 2:
        first_lastname = lst_lastname[0].capitalize()  # делаем первую букву слова заглавной а остальные строчными
        second_lastname = lst_lastname[1].capitalize()
        # Склоняем по отдельности
        first_lastname = maker.make(NamePart.LASTNAME, gender, case, first_lastname)
        second_lastname = maker.make(NamePart.LASTNAME, gender, case, second_lastname)

        return f'{first_lastname}-{second_lastname}'


def case_middlename(maker, middlename, gender, case: Case):
    """
    Функция для обработки и склонения фамилии. Это нужно для обработки случаев двойной фамилии
    """

    lst_middlename = middlename.split('-')  # сплитим по дефису

    if len(lst_middlename) == 1:  # если длина списка равна 1 то это не двойная фамилия и просто обрабатываем слово
        case_result_middlename = maker.make(NamePart.middlename, gender, case, middlename)
        return case_result_middlename
    elif len(lst_middlename) == 2:
        first_middlename = lst_middlename[0].capitalize()  # делаем первую букву слова заглавной а остальные строчными
        second_middlename = lst_middlename[1].capitalize()
        # Склоняем по отдельности
        first_middlename = maker.make(NamePart.MIDDLENAME, gender, case, first_middlename)
        second_middlename = maker.make(NamePart.MIDDLENAME, gender, case, second_middlename)

        return f'{first_middlename}-{second_middlename}'

def detect_gender(detector, lastname, firstname, middlename):
    """
    Функция для определения гендера слова
    """
    #     detector = PetrovichGenderDetector() # создаем объект детектора
    try:
        gender_result = detector.detect(lastname=lastname, firstname=firstname, middlename=middlename)
        return gender_result
    except StopIteration:  # если не удалось определить то считаем что гендер андрогинный
        return Gender.ANDROGYNOUS


def decl_on_case(fio: str, case: Case) -> str:
    """
    Функция для склонения ФИО по падежам
    """
    fio = fio.strip()  # очищаем строку от пробельных символов с начала и конца
    part_fio = fio.split()  # разбиваем по пробелам создавая список где [0] это Фамилия,[1]-Имя,[2]-Отчество

    if len(part_fio) == 3:  # проверяем на длину и обрабатываем только те что имеют длину 3 во всех остальных случаях просим просклонять самостоятельно
        maker = PetrovichDeclinationMaker()  # создаем объект класса
        lastname = part_fio[0].capitalize()  # Фамилия
        firstname = part_fio[1].capitalize()  # Имя
        middlename = part_fio[2].capitalize()  # Отчество

        # Определяем гендер для корректного склонения
        detector = PetrovichGenderDetector()  # создаем объект детектора
        gender = detect_gender(detector, lastname, firstname, middlename)
        # Склоняем

        case_result_lastname = case_lastname(maker, lastname, gender, case)  # обрабатываем фамилию
        case_result_firstname = maker.make(NamePart.FIRSTNAME, gender, case, firstname)
        case_result_firstname = capitalize_double_name(case_result_firstname)  # обрабатываем случаи двойного имени
        case_result_middlename = maker.make(NamePart.MIDDLENAME, gender, case, middlename)
        case_result_middlename = capitalize_double_name(case_result_middlename)
        # Возвращаем результат
        result_fio = f'{case_result_lastname} {case_result_firstname} {case_result_middlename}'
        return result_fio

    else:
        return 'Проверьте количество слов, должно быть 3 разделенных пробелами слова'


def create_initials(cell, checkbox, space):
    """
    Функция для создания инициалов
    """
    lst_fio = cell.split(' ')  # сплитим по пробелу
    if len(lst_fio) == 3:  # проверяем на стандартный размер в 3 слова иначе ничего не меняем
        dct_fio_initials = {'Фамилия': lst_fio[0], 'Имя': lst_fio[1], 'Отчество': lst_fio[2]} # словарь для хранения ицициалов
        for key, value in dct_fio_initials.items():
            if '-' not in value:
                dct_fio_initials[key] = value[0]
            else:
                # если есть дефис то сплитим и добавляем в словарь
                temp = value.split('-')
                first, second = temp[0], temp[1]
                dct_fio_initials[key] = f'{first[0]}-{second[0]}'

        if checkbox == 'ФИ':
            if space == 'без пробела':
                # возвращаем строку вида Иванов И.И.

                return f'{lst_fio[0]} {dct_fio_initials["Имя"].upper()}.{dct_fio_initials["Отчество"].upper()}.'
            else:
                # возвращаем строку с пробелом после имени Иванов И. И.
                return f'{lst_fio[0]} {dct_fio_initials["Имя"].upper()}. {dct_fio_initials["Отчество"].upper()}.'
        else:
            if space == 'без пробела':
                # И.И. Иванов
                return f'{dct_fio_initials["Имя"].upper()}.{dct_fio_initials["Отчество"].upper()}. {lst_fio[0]}'
            else:
                # И. И. Иванов
                return  f'{dct_fio_initials["Имя"].upper()}. {dct_fio_initials["Отчество"].upper()}. {lst_fio[0]}'

    else:
        return cell

def split_fio(value:str,number):
    """
    Функция для разделения данных в колонке ФИО на колонки
    :param value:значение ФИО
    :param number: порядковый номер значения- 0 Фамилия, 1 Имя, 2 Отчество
    :return:
    """
    lst_fio = value.split(' ')
    if len(lst_fio) == 3:
        return lst_fio[number]
    else:
        return 'Проверьте количество слов, должно быть 3 разделенных пробелами слова'


def declension_lst_fio_columns_by_case(df:pd.DataFrame,lst_name_columns:list)->pd.DataFrame:
    """
    Склонение по падежам и создание инициалов по нескольким колонкам с ФИО
    :param df:датафрейм
    :param lst_name_columns: список колонок которые нужно обработать
    :return:измененный датафрейм
    """
    # temp_df = pd.DataFrame()  # временный датафрейм для хранения колонок просклоненных по падежам
    for fio_column in lst_name_columns:
        index_fio_column = lst_name_columns.index(fio_column)  # получаем индекс
        # Обрабатываем nan значения и те которые обозначены пробелом
        df[fio_column].fillna('Не заполнено', inplace=True)
        df[fio_column] = df[fio_column].apply(lambda x: x.strip())
        df[fio_column] = df[fio_column].apply(
            lambda x: x if x else 'Не заполнено')  # Если пустая строка то заменяем на значение Не заполнено
        # создаем колонки
        df[f'{fio_column}_Родительный_падеж'] = df[fio_column].apply(lambda x: decl_on_case(x, Case.GENITIVE))
        df[f'{fio_column}_Дательный_падеж'] = df[fio_column].apply(lambda x: decl_on_case(x, Case.DATIVE))
        df[f'{fio_column}_Винительный_падеж'] = df[fio_column].apply(lambda x: decl_on_case(x, Case.ACCUSATIVE))
        df[f'{fio_column}_Творительный_падеж'] = df[fio_column].apply(lambda x: decl_on_case(x, Case.INSTRUMENTAL))
        df[f'{fio_column}_Предложный_падеж'] = df[fio_column].apply(lambda x: decl_on_case(x, Case.PREPOSITIONAL))
        df[f'{fio_column}_ФИ'] = df[fio_column].apply(lambda x: create_initials(x, 'ФИ', 'без пробела'))
        df[f'{fio_column}_ИФ'] = df[fio_column].apply(lambda x: create_initials(x, 'ИФ', 'без пробела'))
        df[f'{fio_column}_ФИ_пробел'] = df[fio_column].apply(lambda x: create_initials(x, 'ФИ', 'пробел'))
        df[f'{fio_column}_ИФ_пробел'] = df[fio_column].apply(lambda x: create_initials(x, 'ИФ', 'пробел'))

        # Создаем колонки для склонения фамилий с иницалами родительный падеж
        df[f'{fio_column}_ФИ_род_падеж'] = df[f'{fio_column}_Родительный_падеж'].apply(
            lambda x: create_initials(x, 'ФИ', 'без пробела'))
        df[f'{fio_column}_ФИ_род_падеж_пробел'] = df[f'{fio_column}_Родительный_падеж'].apply(
            lambda x: create_initials(x, 'ФИ', 'пробел'))
        df[f'{fio_column}_ИФ_род_падеж'] = df[f'{fio_column}_Родительный_падеж'].apply(
            lambda x: create_initials(x, 'ИФ', 'без пробела'))
        df[f'{fio_column}_ИФ_род_падеж_пробел'] = df[f'{fio_column}_Родительный_падеж'].apply(
            lambda x: create_initials(x, 'ИФ', 'пробел'))

        # Создаем колонки для склонения фамилий с иницалами дательный падеж
        df[f'{fio_column}_ФИ_дат_падеж'] = df[f'{fio_column}_Дательный_падеж'].apply(
            lambda x: create_initials(x, 'ФИ', 'без пробела'))
        df[f'{fio_column}_ФИ_дат_падеж_пробел'] = df[f'{fio_column}_Дательный_падеж'].apply(
            lambda x: create_initials(x, 'ФИ', 'пробел'))
        df[f'{fio_column}_ИФ_дат_падеж'] = df[f'{fio_column}_Дательный_падеж'].apply(
            lambda x: create_initials(x, 'ИФ', 'без пробела'))
        df[f'{fio_column}_ИФ_дат_падеж_пробел'] = df[f'{fio_column}_Дательный_падеж'].apply(
            lambda x: create_initials(x, 'ИФ', 'пробел'))

        # Создаем колонки для склонения фамилий с иницалами винительный падеж
        df[f'{fio_column}_ФИ_вин_падеж'] = df[f'{fio_column}_Винительный_падеж'].apply(
            lambda x: create_initials(x, 'ФИ', 'без пробела'))
        df[f'{fio_column}_ФИ_вин_падеж_пробел'] = df[f'{fio_column}_Винительный_падеж'].apply(
            lambda x: create_initials(x, 'ФИ', 'пробел'))
        df[f'{fio_column}_ИФ_вин_падеж'] = df[f'{fio_column}_Винительный_падеж'].apply(
            lambda x: create_initials(x, 'ИФ', 'без пробела'))
        df[f'{fio_column}_ИФ_вин_падеж_пробел'] = df[f'{fio_column}_Винительный_падеж'].apply(
            lambda x: create_initials(x, 'ИФ', 'пробел'))

        # Создаем колонки для склонения фамилий с иницалами творительный падеж
        df[f'{fio_column}_ФИ_твор_падеж'] = df[f'{fio_column}_Творительный_падеж'].apply(
            lambda x: create_initials(x, 'ФИ', 'без пробела'))
        df[f'{fio_column}_ФИ_твор_падеж_пробел'] = df[f'{fio_column}_Творительный_падеж'].apply(
            lambda x: create_initials(x, 'ФИ', 'пробел'))
        df[f'{fio_column}_ИФ_твор_падеж'] = df[f'{fio_column}_Творительный_падеж'].apply(
            lambda x: create_initials(x, 'ИФ', 'без пробела'))
        df[f'{fio_column}_ИФ_твор_падеж_пробел'] = df[f'{fio_column}_Творительный_падеж'].apply(
            lambda x: create_initials(x, 'ИФ', 'пробел'))
        # Создаем колонки для склонения фамилий с иницалами предложный падеж
        df[f'{fio_column}_ФИ_пред_падеж'] = df[f'{fio_column}_Предложный_падеж'].apply(
            lambda x: create_initials(x, 'ФИ', 'без пробела'))
        df[f'{fio_column}_ФИ_пред_падеж_пробел'] = df[f'{fio_column}_Предложный_падеж'].apply(
            lambda x: create_initials(x, 'ФИ', 'пробел'))
        df[f'{fio_column}_ИФ_пред_падеж'] = df[f'{fio_column}_Предложный_падеж'].apply(
            lambda x: create_initials(x, 'ИФ', 'без пробела'))
        df[f'{fio_column}_ИФ_пред_падеж_пробел'] = df[f'{fio_column}_Предложный_падеж'].apply(
            lambda x: create_initials(x, 'ИФ', 'пробел'))



    return df




def declension_fio_by_case(df:pd.DataFrame,result_folder:str)->pd.DataFrame:
    """
    Функция для склоения фио по падежам , создания инициалов
    :param df: датафрейм с данными
    :param result_folder: конечная папка
    :return: датафрейм
    """
    try:

        temp_df = pd.DataFrame()  # временный датафрейм для хранения колонок просклоненных по падежам
        # проверям одновременное наличие колонок ФИО и Фамилия,Имя,Отчество

        if {'ФИО','Фамилия','Имя','Отчество'}.issubset(set(df.columns)):
            raise FIOTogether
        # проверяем наличие колонки ФИО
        if 'ФИО' in df.columns:
            fio_column = 'ФИО'
            df['ФИО'] = df['ФИО'].apply(lambda x:x.strip() if isinstance(x,str) else 'Не заполнено')
            df['ФИО'] = df['ФИО'].apply(lambda x:re.sub(r'\s+',' ',x))
            # Создаем колонки Фамилия,Имя, Отчество
            df['Фамилия'] = df['ФИО'].apply(lambda x:split_fio(x,0))
            df['Имя'] = df['ФИО'].apply(lambda x:split_fio(x,1))
            df['Отчество'] = df['ФИО'].apply(lambda x:split_fio(x,2))

        else:
            # проверяем наличие колонок
            check_fio_columns = {'Фамилия','Имя','Отчество'}
            diff_cols = check_fio_columns.difference(df.columns)
            if len(diff_cols) != 0:
                raise NotFIOPart
            # Очищаем от пробелов в начале и конце
            df['Фамилия'] = df['Фамилия'].apply(lambda x:x.strip() if isinstance(x,str) else 'Не заполнено')
            df['Имя'] = df['Имя'].apply(lambda x:x.strip() if isinstance(x,str) else 'Не заполнено')
            df['Отчество'] = df['Отчество'].apply(lambda x:x.strip() if isinstance(x,str) else 'Не заполнено')
            df['ФИО'] = df['Фамилия'] + ' ' + df['Имя'] + ' '+ df['Отчество']
            fio_column = 'ФИО'



        # Получаем номер колонки с фио которые нужно обработать
        lst_columns = list(df.columns)  # Превращаем в список
        index_fio_column = lst_columns.index(fio_column)  # получаем индекс

        # Обрабатываем nan значения и те которые обозначены пробелом
        df[fio_column].fillna('Не заполнено', inplace=True)
        df[fio_column] = df[fio_column].apply(lambda x: x.strip())
        df[fio_column] = df[fio_column].apply(
            lambda x: x if x else 'Не заполнено')  # Если пустая строка то заменяем на значение Не заполнено

        temp_df['Родительный_падеж'] = df[fio_column].apply(lambda x: decl_on_case(x, Case.GENITIVE))
        temp_df['Дательный_падеж'] = df[fio_column].apply(lambda x: decl_on_case(x, Case.DATIVE))
        temp_df['Винительный_падеж'] = df[fio_column].apply(lambda x: decl_on_case(x, Case.ACCUSATIVE))
        temp_df['Творительный_падеж'] = df[fio_column].apply(lambda x: decl_on_case(x, Case.INSTRUMENTAL))
        temp_df['Предложный_падеж'] = df[fio_column].apply(lambda x: decl_on_case(x, Case.PREPOSITIONAL))
        temp_df['ФИ'] = df[fio_column].apply(lambda x: create_initials(x, 'ФИ', 'без пробела'))
        temp_df['ИФ'] = df[fio_column].apply(lambda x: create_initials(x, 'ИФ', 'без пробела'))
        temp_df['ФИ_пробел'] = df[fio_column].apply(lambda x: create_initials(x, 'ФИ', 'пробел'))
        temp_df['ИФ_пробел'] = df[fio_column].apply(lambda x: create_initials(x, 'ИФ', 'пробел'))

        # Создаем колонки для склонения фамилий с иницалами родительный падеж
        temp_df['ФИ_род_падеж'] = temp_df['Родительный_падеж'].apply(
            lambda x: create_initials(x, 'ФИ', 'без пробела'))
        temp_df['ФИ_род_падеж_пробел'] = temp_df['Родительный_падеж'].apply(
            lambda x: create_initials(x, 'ФИ', 'пробел'))
        temp_df['ИФ_род_падеж'] = temp_df['Родительный_падеж'].apply(
            lambda x: create_initials(x, 'ИФ', 'без пробела'))
        temp_df['ИФ_род_падеж_пробел'] = temp_df['Родительный_падеж'].apply(
            lambda x: create_initials(x, 'ИФ', 'пробел'))

        # Создаем колонки для склонения фамилий с иницалами дательный падеж
        temp_df['ФИ_дат_падеж'] = temp_df['Дательный_падеж'].apply(
            lambda x: create_initials(x, 'ФИ', 'без пробела'))
        temp_df['ФИ_дат_падеж_пробел'] = temp_df['Дательный_падеж'].apply(
            lambda x: create_initials(x, 'ФИ', 'пробел'))
        temp_df['ИФ_дат_падеж'] = temp_df['Дательный_падеж'].apply(
            lambda x: create_initials(x, 'ИФ', 'без пробела'))
        temp_df['ИФ_дат_падеж_пробел'] = temp_df['Дательный_падеж'].apply(
            lambda x: create_initials(x, 'ИФ', 'пробел'))

        # Создаем колонки для склонения фамилий с иницалами винительный падеж
        temp_df['ФИ_вин_падеж'] = temp_df['Винительный_падеж'].apply(
            lambda x: create_initials(x, 'ФИ', 'без пробела'))
        temp_df['ФИ_вин_падеж_пробел'] = temp_df['Винительный_падеж'].apply(
            lambda x: create_initials(x, 'ФИ', 'пробел'))
        temp_df['ИФ_вин_падеж'] = temp_df['Винительный_падеж'].apply(
            lambda x: create_initials(x, 'ИФ', 'без пробела'))
        temp_df['ИФ_вин_падеж_пробел'] = temp_df['Винительный_падеж'].apply(
            lambda x: create_initials(x, 'ИФ', 'пробел'))

        # Создаем колонки для склонения фамилий с иницалами творительный падеж
        temp_df['ФИ_твор_падеж'] = temp_df['Творительный_падеж'].apply(
            lambda x: create_initials(x, 'ФИ', 'без пробела'))
        temp_df['ФИ_твор_падеж_пробел'] = temp_df['Творительный_падеж'].apply(
            lambda x: create_initials(x, 'ФИ', 'пробел'))
        temp_df['ИФ_твор_падеж'] = temp_df['Творительный_падеж'].apply(
            lambda x: create_initials(x, 'ИФ', 'без пробела'))
        temp_df['ИФ_твор_падеж_пробел'] = temp_df['Творительный_падеж'].apply(
            lambda x: create_initials(x, 'ИФ', 'пробел'))
        # Создаем колонки для склонения фамилий с иницалами предложный падеж
        temp_df['ФИ_пред_падеж'] = temp_df['Предложный_падеж'].apply(
            lambda x: create_initials(x, 'ФИ', 'без пробела'))
        temp_df['ФИ_пред_падеж_пробел'] = temp_df['Предложный_падеж'].apply(
            lambda x: create_initials(x, 'ФИ', 'пробел'))
        temp_df['ИФ_пред_падеж'] = temp_df['Предложный_падеж'].apply(
            lambda x: create_initials(x, 'ИФ', 'без пробела'))
        temp_df['ИФ_пред_падеж_пробел'] = temp_df['Предложный_падеж'].apply(
            lambda x: create_initials(x, 'ИФ', 'пробел'))

        # сохраняем таблицу для проверки правильности склонения
        temp_df.insert(0,'ФИО',df['ФИО'])
        wb = openpyxl.Workbook()
        for row in dataframe_to_rows(temp_df,index=False,header=True):
            wb[wb.sheetnames[0]].append(row)
        for column in wb[wb.sheetnames[0]].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[wb.sheetnames[0]].column_dimensions[column_name].width = adjusted_width
            wb.save(f'{result_folder}/Проверка правильности склонения ФИО слушателей.xlsx')




        # Вставляем получившиеся колонки после базовой колонки с фио
        df.insert(index_fio_column + 1, 'Родительный_падеж', temp_df['Родительный_падеж'])
        df.insert(index_fio_column + 2, 'Дательный_падеж', temp_df['Дательный_падеж'])
        df.insert(index_fio_column + 3, 'Винительный_падеж', temp_df['Винительный_падеж'])
        df.insert(index_fio_column + 4, 'Творительный_падеж', temp_df['Творительный_падеж'])
        df.insert(index_fio_column + 5, 'Предложный_падеж', temp_df['Предложный_падеж'])
        df.insert(index_fio_column + 6, 'ФИ', temp_df['ФИ'])
        df.insert(index_fio_column + 7, 'ИФ', temp_df['ИФ'])
        df.insert(index_fio_column + 8, 'ФИ_пробел', temp_df['ФИ_пробел'])
        df.insert(index_fio_column + 9, 'ИФ_пробел', temp_df['ИФ_пробел'])
        # Добавляем колонки с склонениями инициалов родительный падеж
        df.insert(index_fio_column + 10, 'ФИ_род_падеж', temp_df['ФИ_род_падеж'])
        df.insert(index_fio_column + 11, 'ФИ_род_падеж_пробел',
                  temp_df['ФИ_род_падеж_пробел'])
        df.insert(index_fio_column + 12, 'ИФ_род_падеж', temp_df['ИФ_род_падеж'])
        df.insert(index_fio_column + 13, 'ИФ_род_падеж_пробел',
                  temp_df['ИФ_род_падеж_пробел'])
        # Добавляем колонки с склонениями инициалов дательный падеж
        df.insert(index_fio_column + 14, 'ФИ_дат_падеж', temp_df['ФИ_дат_падеж'])
        df.insert(index_fio_column + 15, 'ФИ_дат_падеж_пробел',
                  temp_df['ФИ_дат_падеж_пробел'])
        df.insert(index_fio_column + 16, 'ИФ_дат_падеж', temp_df['ИФ_дат_падеж'])
        df.insert(index_fio_column + 17, 'ИФ_дат_падеж_пробел',
                  temp_df['ИФ_дат_падеж_пробел'])
        # Добавляем колонки с склонениями инициалов винительный падеж
        df.insert(index_fio_column + 18, 'ФИ_вин_падеж', temp_df['ФИ_вин_падеж'])
        df.insert(index_fio_column + 19, 'ФИ_вин_падеж_пробел',
                  temp_df['ФИ_вин_падеж_пробел'])
        df.insert(index_fio_column + 20, 'ИФ_вин_падеж', temp_df['ИФ_вин_падеж'])
        df.insert(index_fio_column + 21, 'ИФ_вин_падеж_пробел',
                  temp_df['ИФ_вин_падеж_пробел'])
        # Добавляем колонки с склонениями инициалов творительный падеж
        df.insert(index_fio_column + 22, 'ФИ_твор_падеж', temp_df['ФИ_твор_падеж'])
        df.insert(index_fio_column + 23, 'ФИ_твор_падеж_пробел',
                  temp_df['ФИ_твор_падеж_пробел'])
        df.insert(index_fio_column + 24, 'ИФ_твор_падеж', temp_df['ИФ_твор_падеж'])
        df.insert(index_fio_column + 25, 'ИФ_твор_падеж_пробел',
                  temp_df['ИФ_твор_падеж_пробел'])
        # Добавляем колонки с склонениями инициалов предложный падеж
        df.insert(index_fio_column + 26, 'ФИ_пред_падеж', temp_df['ФИ_пред_падеж'])
        df.insert(index_fio_column + 27, 'ФИ_пред_падеж_пробел',
                  temp_df['ФИ_пред_падеж_пробел'])
        df.insert(index_fio_column + 28, 'ИФ_пред_падеж', temp_df['ИФ_пред_падеж'])
        df.insert(index_fio_column + 29, 'ИФ_пред_падеж_пробел',
                  temp_df['ИФ_пред_падеж_пробел'])

        return df

    except NameError:
        messagebox.showerror('Линди Создание документов ДПО,ПО',
                             f'Выберите файлы с данными и папку куда будет генерироваться файл')

    except NotFIOPart:
        messagebox.showerror('Линди Создание документов ДПО,ПО',
                             f'Не найдены колонки Фамилия,Имя,Отчество!')
    except FIOTogether:
        messagebox.showerror('Линди Создание документов ДПО,ПО',
                             f'На листе Данные физлиц одновременно находятся колонки ФИО, Фамилия,Имя, Отчество.\n'
                             f'Удалите колонку ФИО или колонки Фамилия,Имя,Отчество')
    except KeyError as e:
        messagebox.showerror('Линди Создание документов ДПО,ПО',
                             f'В таблице не найдена указанная колонка {e.args}')
    except ValueError:
        messagebox.showerror('Линди Создание документов ДПО,ПО',
                             f'В таблице нет колонки с таким названием!\nПроверьте написание названия колонки')
    except FileNotFoundError:
        messagebox.showerror('Линди Создание документов ДПО,ПО',
                             f'Перенесите файлы, конечную папку с которой вы работете в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам или конечной папке.')

if __name__ == '__main__':
    data_decl_case_main = 'data\Таблица для заполнения бланков.xlsx'
    # data_decl_case_main = 'data\с ФИО.xlsx'
    path_to_end_folder_decl_case_main = 'data/Результат'
    main_df = pd.read_excel(data_decl_case_main,sheet_name='Данные',dtype=str)

    out_df = declension_fio_by_case(main_df)
    out_df.to_excel(f'{path_to_end_folder_decl_case_main}/Результат склонения.xlsx',index=False,header=True)
    print('Lindy Booth')