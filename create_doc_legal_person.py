"""
Скрипт для создания сопроводительной документации для юридических лиц
Основной скрипт
"""
from decl_case import declension_fio_by_case # функция для склонения фио и создания инициалов
from decl_case import declension_lst_fio_columns_by_case # функция для склонения колонок с фио из листа описания курса
from generate_docs import generate_docs # модуль для создания документов
from support_functions import * # вспомогательные функции
import pandas as pd
import numpy as np
import openpyxl
from tkinter import messagebox
import os
from datetime import datetime
import re
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None


class NotNameColumn(Exception):
    """
    Исключение для обработки случая когда не совпадают названия колонок
    """
    pass

class SameNameColumn(Exception):
    """
    Исключение для обработки случая когда в двух листах есть одинаковые названия колонок
    """
    pass

class SamePathFolder(Exception):
    """
    Исключение для случая когда одна и та же папка выбрана в качестве источника и конечной папки
    """
    pass

class NotReqSheet(Exception):
    """
    Исключение для проверки наличия трех листов: Описание, Данные физлиц, Данные юрлиц
    """
    pass
def check_snils(snils):
    """
    Функция для приведения значений снилс в вид ХХХ-ХХХ-ХХХ ХХ
    """
    if snils is np.nan:
        return 'Не заполнено'
    snils = str(snils)
    result = re.findall(r'\d', snils) # ищем цифры
    if len(result) == 11:
        first_group = ''.join(result[:3])
        second_group = ''.join(result[3:6])
        third_group = ''.join(result[6:9])
        four_group = ''.join(result[9:11])

        out_snils = f'{first_group}-{second_group}-{third_group} {four_group}'
        return out_snils
    else:
        return f'Неправильное значение!В СНИЛС физического лица должно быть 11 цифр - {snils} -{len(snils)} цифр'

def create_docs_legal_person(data_file:str,folder_template:str,result_folder:str):
    """
    Скрипт для сопроводительной документации. Точка входа
    :param data_file: файл Excel с данными
    :param folder_template: папка с шаблонами
    :param result_folder: итоговая папка
    :return: Документация в формате docx и файл ФИс-ФРДО
    """
    try:
        if folder_template == result_folder:
            raise SamePathFolder

        # Проверяем наличие листов
        required_sheets = {'Описание','Данные физлиц','Данные юрлиц'}
        req_wb = openpyxl.load_workbook(data_file) # загружаем файл для выяснения состава листов
        diff_sheets = required_sheets.difference(set(req_wb.sheetnames))
        if len(diff_sheets) != 0:
            raise NotReqSheet


        # Предобработка датафрейма с данными курса
        descr_df = pd.read_excel(data_file, sheet_name='Описание', dtype=str,usecols='A:B')  # получаем данные
        descr_df.dropna(how='all',inplace=True) # удаляем пустые строки
        # траснпонируем
        descr_df = descr_df.transpose()
        descr_df.columns = descr_df.iloc[0] # устанавливаем первую строку в качестве названий колонок
        descr_df.drop(labels='Наименование параметра',inplace=True,axis=0) # удаляем первую строку
        descr_df.index = [0] # переименовываем оставшийся индекс в 0
        # Проверяем наличие колонок
        desc_check_cols = {'Наименование_программы','Тип_программы','Вид_документа','Квалификация_профессия_специальность','Разряд_класс','Разряд_класс_текст','Дата_начало','Дата_конец','Объём',
                           'Руководитель','Секретарь','Преподаватель','Куратор','База','Аттестация','Председатель_АК'}
        diff_cols = desc_check_cols.difference(set(descr_df.columns))
        if len(diff_cols) != 0:
            raise NotNameColumn
        descr_df = descr_df.applymap(lambda x:re.sub(r'\s+',' ',x) if isinstance(x,str) else x) # очищаем от лишних пробелов
        descr_df = descr_df.applymap(lambda x:x.strip() if isinstance(x,str) else x) # очищаем от пробелов в начале и конце

        # Получаем тип программы ДПО или ПО
        dpo_set = {'Повышение квалификации','Профессиональная переподготовка'}
        if descr_df.loc[0,'Тип_программы'] in dpo_set:
            type_program = 'ДПО'
        else:
            type_program = 'ПО'


        # Предобработка датафрейма с данными слушателей
        data_df = pd.read_excel(data_file, sheet_name='Данные юрлиц', dtype=str)  # получаем данные
        # Проверяем наличие нужных колонок в файле с данными
        check_columns_data = {'Заказчик','Организация_заказчика','Должность_заказчика','Должность_заказчика_род_падеж','Количество_слушателей',
                              'Стоимость_единицы_ЮЛ','Сумма_договора_кр_ЮЛ'
            ,'Сумма_договора_полн_ЮЛ','Реквизиты_ЮЛ'} # проверяемые колонки
        diff_cols = check_columns_data.difference(set(data_df.columns))
        if len(diff_cols) != 0:
            raise NotNameColumn  # если есть разница вызываем и обрабатываем исключение
        # Обрабатываем колонки из датафрейма с описанием курса склоняя по падежам и создавая иницииалы

        descr_fio_cols =['Заказчик'] # список колонок для которых нужно создать падежи и инициалы
        data_df = declension_lst_fio_columns_by_case(data_df,descr_fio_cols)
        # сохраняем таблицу для проверки правильности склонения
        # получаем список колонок сс словом заказчик чтобы лишние колонки не мешались
        lst_filter = [name_column for name_column in data_df.columns if 'Заказчик' in name_column]
        wb = openpyxl.Workbook()
        for row in dataframe_to_rows(data_df[lst_filter],index=False,header=True):
            wb[wb.sheetnames[0]].append(row)
        for column in wb[wb.sheetnames[0]].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[wb.sheetnames[0]].column_dimensions[column_name].width = adjusted_width
            wb.save(f'{result_folder}/Проверка правильности склонения ФИО заказчиков.xlsx')

        descr_fio_cols =['Руководитель','Секретарь','Преподаватель','Куратор','Председатель_АК'] # список колонок для которых нужно создать падежи и инициалы
        descr_df = declension_lst_fio_columns_by_case(descr_df,descr_fio_cols)

        """
            Конвертируем даты из формата ГГГГ-ММ-ДД в ДД.ММ.ГГГГ
            """
        # делаем строковыми названия колонок
        descr_df.columns = list(map(str,descr_df.columns))
        data_df.columns = list(map(str,data_df.columns))

        # проверяем на совпадение названий колонок в обоих листах
        intersection_columns = set(descr_df.columns).intersection(set(data_df.columns))
        if len(intersection_columns) > 0:
            raise SameNameColumn

        # Обрабатываем колонки с датами в описании
        lst_date_columns_descr = []  # список для колонок с датами
        for idx, column in enumerate(descr_df.columns):
            if 'дата' in column.lower():
                lst_date_columns_descr.append(idx)

        descr_df = convert_string_date(descr_df,lst_date_columns_descr)

        # обрабатываем колонки с датами в списке
        lst_date_columns_data = []  # список для колонок с датами
        for idx, column in enumerate(data_df.columns):
            if 'дата' in column.lower():
                lst_date_columns_data.append(idx)
        data_df = convert_string_date(data_df,lst_date_columns_data)

        # создаем словари с данными для колонок описания программы

        # получаем списки валидных названий колонок
        descr_valid_cols,descr_not_valid_cols = selection_name_column(list(descr_df.columns),r'^[a-zA-ZЁёа-яА-Я_]+$')
        data_valid_cols, data_not_valid_cols = selection_name_column(list(data_df.columns),r'^[a-zA-ZЁёа-яА-Я_]+$')

        # заполняем наны пробелами
        descr_df.fillna(' ',inplace=True)
        data_df.fillna(' ',inplace=True)

        # Словарь с описанием курса
        dct_descr = dict()
        for name_column in descr_valid_cols:
            dct_descr[name_column] = descr_df.loc[0,name_column]

        type_form = 'ЮЛ' # указываем физлицо или юрлицо
        generate_docs(dct_descr,data_df[data_valid_cols],folder_template,result_folder,type_program,type_form)
    except NotNameColumn:
        messagebox.showerror('Линди Создание документов ДПО,ПО',
                             f'В файле {data_file} не найдены следующие колонки {diff_cols}')
    except SameNameColumn:
        messagebox.showerror('Линди Создание документов ДПО,ПО',
                             f'На листе с описанием и на листе со списком найдены одинаковые названия колонок {intersection_columns}\n'
                             f'переименуйте колонки')

    except SamePathFolder:
        messagebox.showerror('Линди Создание документов ДПО,ПО',
                             f'Выбрана одна и та же папка в качесте исходной и конечной.\n'
                             f'Исходная и конечная папки должны быть разными !!!')
    except NotReqSheet:
        messagebox.showerror('Линди Создание документов ДПО,ПО',
                             f'В файле с данными курса не найдены обязательные листы {diff_sheets}')
    except PermissionError as e:
        messagebox.showerror('Линди Создание документов ДПО,ПО',
                             f'Закройте файлы созданные программой')
    else:
        messagebox.showinfo('Линди Создание документов ДПО,ПО', 'Создание документов успешно завершено !')



if __name__ == '__main__':
    main_data_file = 'data/Организация ключевых процессов для повышения эффективности.xlsx'
    main_folder_template = 'data/Шаблоны'
    # main_folder_template = 'data/Шаблоны/empty'
    main_result_folder = 'data/Результат'

    create_docs_legal_person(main_data_file,main_folder_template,main_result_folder)
    print('Lindy Booth !!!')