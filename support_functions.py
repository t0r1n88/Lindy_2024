"""
Вспомогательные функции
"""
import datetime
import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import re

def convert_date_yandex(value:str):
    """
    Функция для конвертации дат из яндекс файла
    :param value: строка
    :return:дата
    """
    try:
        date_object = datetime.datetime.strptime(value, "%Y-%m-%d") # делаем объект datetime
        return date_object.strftime("%d.%m.%Y") # преобразуем в нужный формат
    except ValueError:
        return value

def create_doc_convert_date(cell):
    """
    Функция для конвертации даты при создании документов
    :param cell:
    :return:
    """
    try:
        string_date = datetime.datetime.strftime(cell, '%d.%m.%Y')
        return string_date
    except ValueError:
        return ''
    except TypeError:
        return ''


def convert_string_date(df:pd.DataFrame,lst_date_columns:list)->pd.DataFrame:
    """
    Функция для коневертации колонок с датами в строковый формат для правильного отображения
    :param df: датафрейм с данными
    :param lst_date_columns: список с индексами колонок с датами
    :return: исправленный датафрейм
    """
    lst_name_columns = [] # список куда будут сохраняться названия колонок
    for i in lst_date_columns:
        lst_name_columns.append(list(df.columns)[i])

    # Конвертируем в пригодный строковый формат
    for name_column in lst_name_columns:
        df[name_column] = pd.to_datetime(df[name_column],errors='coerce')
        df[name_column] =  df[name_column].apply(create_doc_convert_date)


    return df



def write_df_to_excel(dct_df:dict,write_index:bool)->openpyxl.Workbook:
    """
    Функция для записи датафрейма в файл Excel
    :param dct_df: словарь где ключе это название создаваемого листа а значение датафрейм который нужно записать
    :param write_index: нужно ли записывать индекс датафрейма True or False
    :return: объект Workbook с записанными датафреймами
    """
    wb = openpyxl.Workbook() # создаем файл
    count_index = 0 # счетчик индексов создаваемых листов
    for name_sheet,df in dct_df.items():
        wb.create_sheet(title=name_sheet,index=count_index) # создаем лист
        # записываем данные в лист
        for row in dataframe_to_rows(df,index=write_index,header=True):
            wb[name_sheet].append(row)
        # ширина по содержимому
        # сохраняем по ширине колонок
        for column in wb[name_sheet].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[name_sheet].column_dimensions[column_name].width = adjusted_width
        count_index += 1
    # удаляем лишний лист
    if len(wb.sheetnames) >= 2 and 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    return wb


def selection_name_column(lst_cols: list, pattern: str):
    """
    Функция для отбора значений попадающих под условие
    :param lst_cols: список с строками
    :param pattern: паттерн отбора
    :return:кортеж из 2 списков, первй список это подошедшие под условие а второй список это не подошедшие
    """
    valid_cols = [name_col for name_col in lst_cols if re.search(pattern,name_col)]
    not_valid_cols = (set(lst_cols)).difference(set(valid_cols))
    return valid_cols,not_valid_cols




