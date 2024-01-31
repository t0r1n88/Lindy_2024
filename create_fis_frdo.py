"""
Модуль для создания файла ФИС ФРДО
"""
from support_functions import * # вспомогательные функции
import re
import pandas as pd
import openpyxl
from tkinter import messagebox
import os
import time
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None


class NotFileTemplateDPO(Exception):
    """
    Класс для ошибки когда не найден файл шаблона ДПО
    """
    pass

class NotFileTemplatePO(Exception):
    """
    Класс для ошибки когда не найден файл шаблона ПО
    """
    pass

class NotNameColumn(Exception):
    """
    Исключение для обработки случая когда не совпадают названия колонок
    """
    pass

def write_data_fis_frdo(template_fis_frdo_dpo:openpyxl.Workbook,dct_df:dict,dct_number_column:dict,dct_descr_number:dict,dct_descr_df:dict,
                        dct_constant_number_column:dict,dct_constant_value_column:dict)->openpyxl.Workbook:
    """
    Функция для записи данных в шаблон ФИС -ФРДО
    :param template_fis_frdo_dpo: шаблон ФИС-ФРДО
    :param dct_df: словарь с данными вида -название колонки:список данных в колонке
    :param dct_number_column: словарь вида -название колонки: порядковый номер колонки куда надо записывать данные
    :param dct_descr_number: словарь вида -название колонки: порядковый номер колонки куда надо записывать данные для описания курса
    :param dct_descr_df: датафрейм с описанием программы
    :param dct_constant_number_column: словарь с номерами колонок для константных значений
    :param dct_constant_value_column: словарь с значениями констант
    :return: заполненый шаблон
    """
    count_row = len(dct_df['Фамилия']) # получаем количество строк в датафрейме
    # извлекаем год
    date_begin = dct_descr_df['Дата_начало'][0]
    date_end = dct_descr_df['Дата_конец'][0]
    result = re.search(r'\d{4}',date_begin)
    if result:
        date_begin = int(result.group())
    else:
        date_begin = 'Не найден год в формате ГГГГ. Проверьте правильность написания'

    result = re.search(r'\d{4}',date_end)
    if result:
        date_end = int(result.group())
    else:
        date_end = 'Не найден год в формате ГГГГ. Проверьте правильность написания'

    dct_descr_df['Дата_начало'][0] = date_begin
    dct_descr_df['Дата_конец'][0] = date_end
    # Делаем числовым срок обучения
    temp_volume =str(dct_descr_df['Объём'][0])
    if temp_volume.isdigit():
        dct_descr_df['Объём'][0] = int(temp_volume)
    else:
        dct_descr_df['Объём'][0] = 'Некорректное значение. Должно быть указано целое число'

    for name_column,number_col in dct_number_column.items():
        # перебираем словарь с порядковыми номерами колонок
        start_row = 2  # строка с которой будет начинаться записи
        for value in dct_df[name_column]: # записываем данные из словаря с данными
            template_fis_frdo_dpo['Шаблон'].cell(row=start_row, column=number_col, value=value)
            start_row += 1

    # пха тройный цикл ну да ладно
    for name_column,number_col in dct_descr_number.items():
        for value in dct_descr_df[name_column]:
            for row in range(2,count_row+2): # записываем данные из словаря с описанием
                template_fis_frdo_dpo['Шаблон'].cell(row=row, column=number_col, value=value)


    for name_column,number_col in dct_constant_number_column.items():
        for row in range(2,count_row+2): # записываем данные из словаря с константами
            template_fis_frdo_dpo['Шаблон'].cell(row=row, column=number_col, value=dct_constant_value_column[name_column])

    return template_fis_frdo_dpo



def create_fis_frdo(df:pd.DataFrame,descr_df:pd.DataFrame,folder_template:str,result_folder:str,type_program:str,type_doc:str):
    """
    Функция для создания файлов ФИС ФРДО
    :param df: датафрейм с данными
    :param descr_df: датафрейм с описанием курса
    :param result_folder: путь к конечной папке
    :param folder_template:путь к папке с шаблонами
    :param type_program: тип создаваемого файла - ПК или ПО
    :param type_doc: вид создаваемого документа
    :return:файл Excel
    """
    # генерируем текущее время
    t = time.localtime()
    current_time = time.strftime('%H_%M_%S', t)
    try:
        if type_program == 'ДПО':
            dct_df = df.to_dict(orient='list') # превращаем в словарь где ключ это название колонки а значение это список
            dct_descr_df = descr_df.to_dict(orient='list')
            # Создаем словарь для хранения номеров колонок для каждого названия
            dct_number_column = {'Номер_удостоверения':7,'Рег_номер':9,'Уровень_образования':15,'Фамилия_диплом':16,'Серия_диплом':17,'Номер_диплом':18,
                                 'Фамилия':22,
                                 'Имя':23,'Отчество':24,
                                 'Дата_рождения':25,'Пол':26,'СНИЛС':27}

            dct_descr_number = {'Тип_программы':10,'Наименование_программы':11,'Квалификация_профессия_специальность':14,
                                 'Дата_начало':19,'Дата_конец':20,'Объём':21}

            # словари для данных которые не меняются
            dct_constant_number_column = {'Вид_документа':1,'Статус документа':2,'Подтверждение утраты':3,'Подтверждение обмена':4,'Подтверждение уничтожения':5,
            'Серия документа':6,'Форма получения образования на момент прекращения образовательных отношений':30}

            dct_constant_value_column = {'Вид_документа': type_doc,'Статус документа':'Оригинал', 'Подтверждение утраты':'Нет', 'Подтверждение обмена':'Нет', 'Подтверждение уничтожения':'Нет',
            'Серия документа':'нет', 'Форма получения образования на момент прекращения образовательных отношений':'в образовательной организации'}



            template_fis_frdo_dpo = openpyxl.load_workbook(f'{folder_template}/ФИС-ФРДО/Шаблон ФИС-ФРДО ДПО.xlsx')
            fis_frdo_dpo = write_data_fis_frdo(template_fis_frdo_dpo,dct_df,dct_number_column,dct_descr_number,dct_descr_df,
                                               dct_constant_number_column,dct_constant_value_column) # Записываем в шаблон
            # делаем колонки по ширине содержимого
            for column in fis_frdo_dpo['Шаблон'].columns:
                max_length = 0
                column_name = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                fis_frdo_dpo['Шаблон'].column_dimensions[column_name].width = adjusted_width
                # Изменяем у некоторых колонок ширину
                fis_frdo_dpo['Шаблон'].column_dimensions['J'].width = 50
                fis_frdo_dpo['Шаблон'].column_dimensions['S'].width = 20
                fis_frdo_dpo['Шаблон'].column_dimensions['T'].width = 20
                fis_frdo_dpo['Шаблон'].column_dimensions['U'].width = 20
                fis_frdo_dpo['Шаблон'].column_dimensions['AD'].width = 30
                fis_frdo_dpo['Шаблон'].column_dimensions['AE'].width = 20

            # проверяем наличие папки ФИС-ФРДО
            if not os.path.exists(f'{result_folder}/ФИС-ФРДО'):
                os.makedirs(f'{result_folder}/ФИС-ФРДО')
            fis_frdo_dpo.save(f'{result_folder}/ФИС-ФРДО/ФИС-ФРДО ДПО {current_time}.xlsx')
        elif type_program == 'ПО':
            dct_df = df.to_dict(
                orient='list')  # превращаем в словарь где ключ это название колонки а значение это список
            dct_descr_df = descr_df.to_dict(orient='list')
            # Создаем словарь для хранения номеров колонок для каждого названия
            dct_number_column = {'Номер_удостоверения':7,'Рег_номер':9,
                                 'Фамилия':17,
                                 'Имя':18,'Отчество':19,
                                 'Дата_рождения':20,'Пол':21,'СНИЛС':22}

            dct_descr_number = {'Тип_программы':10,'Наименование_программы':11,'Квалификация_профессия_специальность':12,'Разряд_класс':13,
                                 'Дата_начало':14,'Дата_конец':15,'Объём':16}

            # словари для данных которые не меняются
            dct_constant_number_column = {'Вид_документа':1,'Статус документа':2,'Подтверждение утраты':3,'Подтверждение обмена':4,'Подтверждение уничтожения':5,
            'Серия документа':6,'Форма получения образования на момент прекращения образовательных отношений':26}

            dct_constant_value_column = {'Вид_документа':type_doc,'Статус документа':'Оригинал', 'Подтверждение утраты':'Нет', 'Подтверждение обмена':'Нет', 'Подтверждение уничтожения':'Нет',
            'Серия документа':'нет', 'Форма получения образования на момент прекращения образовательных отношений':'в образовательной организации'}

            template_fis_frdo_po = openpyxl.load_workbook(f'{folder_template}/ФИС-ФРДО/Шаблон ФИС-ФРДО ПО.xlsx')
            fis_frdo_po = write_data_fis_frdo(template_fis_frdo_po, dct_df, dct_number_column, dct_descr_number,
                                               dct_descr_df,dct_constant_number_column,dct_constant_value_column)  # Записываем в шаблон
            # делаем колонки по ширине содержимого
            for column in fis_frdo_po['Шаблон'].columns:
                max_length = 0
                column_name = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                fis_frdo_po['Шаблон'].column_dimensions[column_name].width = adjusted_width
                # Изменяем у некоторых колонок ширину
                fis_frdo_po['Шаблон'].column_dimensions['J'].width = 50
                fis_frdo_po['Шаблон'].column_dimensions['K'].width = 50
                fis_frdo_po['Шаблон'].column_dimensions['L'].width = 70
                fis_frdo_po['Шаблон'].column_dimensions['M'].width = 20
                fis_frdo_po['Шаблон'].column_dimensions['S'].width = 20
                fis_frdo_po['Шаблон'].column_dimensions['T'].width = 20
                fis_frdo_po['Шаблон'].column_dimensions['U'].width = 20
                fis_frdo_po['Шаблон'].column_dimensions['W'].width = 20
                fis_frdo_po['Шаблон'].column_dimensions['Z'].width = 30
                fis_frdo_po['Шаблон'].column_dimensions['AD'].width = 30
                fis_frdo_po['Шаблон'].column_dimensions['AE'].width = 20
            # проверяем наличие папки ФИС-ФРДО
            if not os.path.exists(f'{result_folder}/ФИС-ФРДО'):
                os.makedirs(f'{result_folder}/ФИС-ФРДО')
            fis_frdo_po.save(f'{result_folder}/ФИС-ФРДО/ФИС-ФРДО ПО {current_time}.xlsx')



    except FileNotFoundError:
        messagebox.showerror('Создание документов ДПО,ПО',
                             f'В папке {folder_template}/ФИС-ФРДО не найден файл шаблона ФИС-ФРДО.\n'
                                                              f'Файлы должны иметь название - Шаблон ФИС-ФРДО ДПО и Шаблон ФИС-ФРДО ПО')






if __name__ == '__main__':
    main_data_file = 'data/Таблица для заполнения бланков.xlsx'
    main_folder_template = 'data/Шаблоны'
    main_result_folder = 'data/Результат'
    main_type_program = 'ДПО'
    # Предобработка датафрейма с данными курса
    descr_df = pd.read_excel(main_data_file, sheet_name='Описание', dtype=str, nrows=1)  # получаем данные
    # Проверяем наличие колонок
    desc_check_cols = {'Наименование_программы', 'Тип_программы', 'Квалификация', 'Дата_начало', 'Дата_конец', 'Объем',
                       'ФИО_руководитель', 'Должность_руководитель', 'Основание_родит_падеж', 'ФИО_секретарь', 'База'}
    diff_cols = desc_check_cols.difference(set(descr_df.columns))
    if len(diff_cols) != 0:
        raise NotNameColumn
    descr_df = descr_df.applymap(
        lambda x: re.sub(r'\s+', ' ', x) if isinstance(x, str) else x)  # очищаем от лишних пробелов
    descr_df = descr_df.applymap(
        lambda x: x.strip() if isinstance(x, str) else x)  # очищаем от пробелов в начале и конце

    # Создаем единичные переменные
    name_program = descr_df.loc[0, 'Наименование_программы']
    type_course = descr_df.loc[0, 'Тип_программы']
    name_qval = descr_df.loc[0, 'Квалификация']
    date_begin = descr_df.loc[0, 'Дата_начало']
    date_end = descr_df.loc[0, 'Дата_конец']
    volume = descr_df.loc[0, 'Объем']
    fio_chief = descr_df.loc[0, 'ФИО_руководитель']
    chief_position = descr_df.loc[0, 'Должность_руководитель']
    name_doc_rod_case = descr_df.loc[0, 'Основание_родит_падеж']
    fio_secretary = descr_df.loc[0, 'ФИО_секретарь']
    base = descr_df.loc[0, 'База']

    # Предобработка датафрейма с данными слушателей
    data_df = pd.read_excel(main_data_file, sheet_name='Данные', dtype=str)  # получаем данные
    # Проверяем наличие нужных колонок в файле с данными
    check_columns_data = {'Номер_удостоверения', 'Рег_номер', 'Дата_рождения', 'Пол', 'СНИЛС', 'Гражданство',
                          'Уровень_образования'
        , 'Серия_паспорта', 'Номер_паспорта', 'Кем_выдан_паспорт', 'Дата_выдачи_паспорта'}  # проверяемые колонки
    diff_cols = check_columns_data.difference(set(data_df.columns))
    if len(diff_cols) != 0:
        raise NotNameColumn  # если есть разница вызываем и обрабатываем исключение
    """
        Конвертируем даты из формата ГГГГ-ММ-ДД в ДД.ММ.ГГГГ
        """
    data_df['Дата_рождения'] = data_df['Дата_рождения'].apply(convert_date_yandex)
    data_df['Дата_выдачи_паспорта'] = data_df['Дата_выдачи_паспорта'].apply(convert_date_yandex)

    create_fis_frdo(data_df,descr_df,main_folder_template,main_result_folder,main_type_program,main_data_file)
    print('Lindy Booth !!!')

